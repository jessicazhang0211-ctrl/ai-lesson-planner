from flask import Blueprint, request, Response, make_response
from flask import current_app
from app.utils.response import ok, err
from app.extensions import db
from app.models.classroom import Classroom, Student
from app.models.resource_publish import ResourcePublish
from app.models.resource_assignment import ResourceAssignment
import datetime, random, string
import io, csv
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except Exception:
    openpyxl = None
try:
    import xlrd
except Exception:
    xlrd = None

bp = Blueprint("class", __name__, url_prefix="/api/class")


def _gen_code(n=6):
    return ''.join(random.choices(string.ascii_uppercase + string.digits, k=n))


def _normalize_header(h):
    """Normalize a header cell: strip whitespace, remove BOM, lower-case."""
    if h is None:
        return ''
    try:
        s = str(h)
    except Exception:
        s = ''
    return s.lstrip('\ufeff').strip().lower()


def _find_header_row(rows):
    """Find the first row that contains required header columns (name/stu_id or 姓名/学号)."""
    max_scan = min(len(rows), 8)
    for i in range(max_scan):
        norm = [_normalize_header(h) for h in rows[i]]
        hk = set(norm)
        if (('name' in hk or '姓名' in hk) and ('stu_id' in hk or '学号' in hk)):
            return norm, i
    return None, None


def _debug_headers_allowed():
    # allow when app is in debug or when request explicitly asks for header debug
    try:
        if getattr(current_app, 'debug', False):
            return True
    except Exception:
        pass
    v = (request.args.get('debug_headers') or '').lower()
    return v in ('1', 'true', 'yes')


def _get_uid():
    uid = (request.headers.get("X-User-Id") or "").strip()
    try:
        return int(uid) if uid else None
    except ValueError:
        return None


@bp.route("/", methods=["GET"])
def list_classes():
    uid = _get_uid()
    if not uid:
        return err("missing X-User-Id", http_status=401)

    status = request.args.get('status')  # optional filter
    q = Classroom.query.filter_by(created_by=uid)
    if status in ("active", "archived"):
        q = q.filter_by(status=status)
    classes = q.order_by(Classroom.created_at.desc()).all()
    return ok([c.to_dict() for c in classes])


@bp.route("/public", methods=["GET"])
def list_public_classes():
    classes = Classroom.query.filter_by(status="active").order_by(Classroom.created_at.desc()).all()
    return ok([{"id": c.id, "name": c.name} for c in classes])


@bp.route("/", methods=["POST"])
def create_class():
    uid = _get_uid()
    if not uid:
        return err("missing X-User-Id", http_status=401)

    data = request.get_json(silent=True) or {}
    name = (data.get('name') or '').strip()
    if not name:
        return err('name required', http_status=400)

    c = Classroom(
        name=name,
        description=data.get('desc') or '',
        status='active',
        code=_gen_code(6),
        created_at=datetime.datetime.now(),
        created_by=uid,
        stage=data.get('stage') or '',
        allow_join=bool(data.get('allow_join', True)),
        note=data.get('note') or ''
    )
    db.session.add(c)
    db.session.commit()
    return ok(c.to_dict())


@bp.route('/<int:cid>', methods=['GET'])
def get_class(cid: int):
    uid = _get_uid()
    if not uid:
        return err('missing X-User-Id', http_status=401)
    c = Classroom.query.get(cid)
    if not c or c.created_by != uid:
        return err('class not found', http_status=404)
    return ok(c.to_dict(include_students=True))


@bp.route('/<int:cid>/stats/basic', methods=['GET'])
def class_stats_basic(cid: int):
    uid = _get_uid()
    if not uid:
        return err('missing X-User-Id', http_status=401)
    c = Classroom.query.get(cid)
    if not c or c.created_by != uid:
        return err('class not found', http_status=404)

    pubs = ResourcePublish.query.filter_by(class_id=cid, revoked=False, resource_type='exercise').all()
    pub_ids = [p.id for p in pubs]
    assignments = ResourceAssignment.query.filter(ResourceAssignment.publish_id.in_(pub_ids)).all() if pub_ids else []

    total = len(assignments)
    completed = len([a for a in assignments if a.status == 'completed'])
    submission_rate = int(round((completed / total) * 100)) if total else 0

    scores = [a.score for a in assignments if a.score is not None]
    avg_score = int(round(sum(scores) / len(scores))) if scores else None

    return ok({
        'class_id': cid,
        'submission_rate': submission_rate,
        'completed': completed,
        'total': total,
        'avg_score': avg_score
    })


@bp.route('/<int:cid>', methods=['PATCH'])
def update_class(cid: int):
    uid = _get_uid()
    if not uid:
        return err('missing X-User-Id', http_status=401)
    c = Classroom.query.get(cid)
    if not c or c.created_by != uid:
        return err('class not found', http_status=404)

    data = request.get_json(silent=True) or {}
    allowed = {'name', 'desc', 'stage', 'allow_join', 'note'}
    for k in allowed:
        if k in data:
            if k == 'desc':
                setattr(c, 'description', data.get(k))
            else:
                setattr(c, k, data.get(k))
    db.session.commit()
    return ok(c.to_dict())


@bp.route('/<int:cid>/archive', methods=['POST'])
def archive_class(cid: int):
    uid = _get_uid()
    if not uid:
        return err('missing X-User-Id', http_status=401)
    c = Classroom.query.get(cid)
    if not c or c.created_by != uid:
        return err('class not found', http_status=404)
    data = request.get_json(silent=True) or {}
    action = data.get('action', 'archive')
    if action == 'archive':
        c.status = 'archived'
    else:
        c.status = 'active'
    db.session.commit()
    return ok({'status': c.status})


@bp.route('/<int:cid>', methods=['DELETE'])
def delete_class(cid: int):
    uid = _get_uid()
    if not uid:
        return err('missing X-User-Id', http_status=401)
    c = Classroom.query.get(cid)
    if not c or c.created_by != uid:
        return err('class not found', http_status=404)
    db.session.delete(c)
    db.session.commit()
    return ok({'deleted': True}, 'deleted')


# Students
@bp.route('/<int:cid>/students', methods=['POST'])
def add_student(cid: int):
    uid = _get_uid()
    if not uid:
        return err('missing X-User-Id', http_status=401)
    c = Classroom.query.get(cid)
    if not c or c.created_by != uid:
        return err('class not found', http_status=404)
    data = request.get_json(silent=True) or {}
    name = (data.get('name') or '').strip()
    if not name:
        return err('student name required', http_status=400)
    s = Student(
        name=name,
        stu_id=data.get('stu_id') or '',
        status=data.get('status') or 'joined',
        parent_phone=data.get('parent_phone') or '',
        accuracy=data.get('accuracy'),
        submit=data.get('submit'),
        class_id=cid
    )
    db.session.add(s)
    db.session.commit()
    return ok(s.to_dict())


@bp.route('/join', methods=['POST', 'OPTIONS'])
def join_by_code():
    if request.method == 'OPTIONS':
        return '', 204

    data = request.get_json(silent=True) or {}
    code = (data.get('code') or '').strip()
    name = (data.get('name') or '').strip()
    stu_id = (data.get('stu_id') or '').strip()
    parent_phone = (data.get('parent_phone') or '').strip()

    if not code or not name:
        return err('code and name required', http_status=400)

    cls = Classroom.query.filter_by(code=code, status='active').first()
    if not cls:
        return err('class not found or not joinable', http_status=404)
    if not cls.allow_join:
        return err('class does not allow join', http_status=403)

    s = Student(
        name=name,
        stu_id=stu_id or '',
        status='joined',
        parent_phone=parent_phone or '',
        class_id=cls.id
    )
    db.session.add(s)
    db.session.commit()

    return ok({'class': cls.to_dict(), 'student': s.to_dict()}, 'joined')




@bp.route('/<int:cid>/students/<int:sid>', methods=['PATCH'])
def update_student(cid: int, sid: int):
    uid = _get_uid()
    if not uid:
        return err('missing X-User-Id', http_status=401)
    c = Classroom.query.get(cid)
    if not c or c.created_by != uid:
        return err('class not found', http_status=404)
    s = Student.query.get(sid)
    if not s or s.class_id != cid:
        return err('student not found', http_status=404)
    data = request.get_json(silent=True) or {}
    for k in ('name', 'stu_id', 'status', 'parent_phone', 'accuracy', 'submit'):
        if k in data:
            setattr(s, k if k != 'parent_phone' else 'parent_phone', data.get(k))
    db.session.commit()
    return ok(s.to_dict())


@bp.route('/<int:cid>/students/<int:sid>', methods=['DELETE'])
def delete_student(cid: int, sid: int):
    uid = _get_uid()
    if not uid:
        return err('missing X-User-Id', http_status=401)
    c = Classroom.query.get(cid)
    if not c or c.created_by != uid:
        return err('class not found', http_status=404)
    s = Student.query.get(sid)
    if not s or s.class_id != cid:
        return err('student not found', http_status=404)
    db.session.delete(s)
    db.session.commit()
    return ok({'deleted': True}, 'deleted')


@bp.route('/<int:cid>/students/<int:sid>', methods=['GET'])
def get_student(cid: int, sid: int):
    uid = _get_uid()
    if not uid:
        return err('missing X-User-Id', http_status=401)
    c = Classroom.query.get(cid)
    if not c or c.created_by != uid:
        return err('class not found', http_status=404)
    s = Student.query.get(sid)
    if not s or s.class_id != cid:
        return err('student not found', http_status=404)
    return ok(s.to_dict())


def _gen_temp_password(n=8):
    return ''.join(random.choices(string.ascii_letters + string.digits, k=n))


@bp.route('/<int:cid>/students/<int:sid>/reset-password', methods=['POST'])
def reset_student_password(cid: int, sid: int):
    """Demo endpoint: generate a temporary password for the student and return it.
    In a real system this would integrate with the auth/user table and send via SMS/email.
    """
    uid = _get_uid()
    if not uid:
        return err('missing X-User-Id', http_status=401)
    c = Classroom.query.get(cid)
    if not c or c.created_by != uid:
        return err('class not found', http_status=404)
    s = Student.query.get(sid)
    if not s or s.class_id != cid:
        return err('student not found', http_status=404)

    newpwd = _gen_temp_password(8)
    # Note: Student model doesn't store passwords. This is a demo-only response.
    return ok({'new_password': newpwd}, 'password reset (demo)')


@bp.route('/<int:cid>/students/<int:sid>/status', methods=['POST'])
def set_student_status(cid: int, sid: int):
    uid = _get_uid()
    if not uid:
        return err('missing X-User-Id', http_status=401)
    c = Classroom.query.get(cid)
    if not c or c.created_by != uid:
        return err('class not found', http_status=404)
    s = Student.query.get(sid)
    if not s or s.class_id != cid:
        return err('student not found', http_status=404)
    data = request.get_json(silent=True) or {}
    action = (data.get('action') or '').lower()
    if action == 'disable':
        s.status = 'disabled'
    elif action == 'enable':
        s.status = 'joined'
    else:
        return err('invalid action', http_status=400)
    db.session.commit()
    return ok({'status': s.status})


@bp.route('/<int:cid>/reset-code', methods=['POST'])
def reset_class_code(cid: int):
    uid = _get_uid()
    if not uid:
        return err('missing X-User-Id', http_status=401)
    c = Classroom.query.get(cid)
    if not c or c.created_by != uid:
        return err('class not found', http_status=404)
    c.code = _gen_code(6)
    db.session.commit()
    return ok({'code': c.code}, 'code reset')


@bp.route('/<int:cid>/export', methods=['GET'])
def export_class(cid: int):
    uid = _get_uid()
    if not uid:
        return err('missing X-User-Id', http_status=401)
    c = Classroom.query.get(cid)
    if not c or c.created_by != uid:
        return err('class not found', http_status=404)

    fmt = (request.args.get('format') or '').lower()
    students = c.students or []
    if fmt == 'csv':
        si = io.StringIO()
        writer = csv.writer(si)
        writer.writerow(['name', 'stu_id', 'status', 'parent_phone', 'accuracy', 'submit', 'created_at'])
        for s in students:
            writer.writerow([s.name or '', s.stu_id or '', s.status or '', s.parent_phone or '', s.accuracy if s.accuracy is not None else '', s.submit if s.submit is not None else '', s.created_at.strftime('%Y-%m-%d %H:%M:%S') if s.created_at else ''])
        output = si.getvalue()
        resp = make_response(output)
        resp.headers['Content-Type'] = 'text/csv; charset=utf-8'
        resp.headers['Content-Disposition'] = f'attachment; filename=class_{c.id}_students.csv'
        return resp
    if fmt in ('xlsx', 'excel'):
        if not openpyxl:
            return err('xlsx export requires openpyxl', http_status=500)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Students'
        headers = ['name', 'stu_id', 'status', 'parent_phone', 'accuracy', 'submit', 'created_at']
        ws.append(headers)
        for s in students:
            ws.append([s.name or '', s.stu_id or '', s.status or '', s.parent_phone or '', s.accuracy if s.accuracy is not None else '', s.submit if s.submit is not None else '', s.created_at.strftime('%Y-%m-%d %H:%M:%S') if s.created_at else ''])
        # auto width
        for i, col in enumerate(ws.columns, 1):
            max_length = 0
            for cell in col:
                try:
                    v = str(cell.value or '')
                except Exception:
                    v = ''
                if len(v) > max_length:
                    max_length = len(v)
            ws.column_dimensions[get_column_letter(i)].width = min(50, max(10, max_length + 2))
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        resp = make_response(bio.read())
        resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        resp.headers['Content-Disposition'] = f'attachment; filename=class_{c.id}_students.xlsx'
        return resp

    # default: return JSON with students
    return ok(c.to_dict(include_students=True))


@bp.route('/<int:cid>/import', methods=['POST'])
def import_students(cid: int):
    """Accepts JSON body with {'students': [...]} or raw CSV in request data or form field 'csv'.
    Matching is attempted by stu_id (when provided) to update existing students; otherwise new students are created.
    """
    uid = _get_uid()
    if not uid:
        return err('missing X-User-Id', http_status=401)
    c = Classroom.query.get(cid)
    if not c or c.created_by != uid:
        return err('class not found', http_status=404)

    added = 0
    updated = 0

    data = request.get_json(silent=True)
    # case 1: JSON payload with students array
    if data and isinstance(data, dict) and 'students' in data and isinstance(data['students'], list):
        rows = data['students']
    else:
        # case 2: file upload via multipart/form-data
        if 'file' in request.files:
            f = request.files['file']
            filename = (f.filename or '').lower()
            content = f.read()
            # xlsx
            if filename.endswith('.xlsx') or (openpyxl and content[:2] == b'PK'):
                if not openpyxl:
                    return err('xlsx import requires openpyxl', http_status=500)
                wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
                ws = wb.active
                rows_iter = list(ws.iter_rows(values_only=True))
                if not rows_iter:
                    return err('empty xlsx', http_status=400)
                header, header_idx = _find_header_row(rows_iter)
                if header is None:
                    data = {'parsed_header': rows_iter[0]} if _debug_headers_allowed() else None
                    return err('invalid header: expected name and stu_id (or 姓名 and 学号)', http_status=400, data=data)
                # trim leading non-header rows
                if header_idx and header_idx > 0:
                    rows_iter = rows_iter[header_idx:]
                # map indexes
                def find_idx(keys):
                    for k in keys:
                        if k in header:
                            return header.index(k)
                    return None
                idx_name = find_idx(['name','姓名'])
                idx_sid = find_idx(['stu_id','学号'])
                idx_phone = find_idx(['parent_phone','家长电话'])
                idx_status = find_idx(['status','状态'])
                rows = []
                for row in rows_iter[1:]:
                    name = str(row[idx_name]).strip() if (idx_name is not None and idx_name < len(row) and row[idx_name] is not None) else ''
                    sid = str(row[idx_sid]).strip() if (idx_sid is not None and idx_sid < len(row) and row[idx_sid] is not None) else ''
                    phone = str(row[idx_phone]).strip() if (idx_phone is not None and idx_phone < len(row) and row[idx_phone] is not None) else ''
                    status = str(row[idx_status]).strip() if (idx_status is not None and idx_status < len(row) and row[idx_status] is not None) else 'joined'
                    if not name:
                        continue
                    rows.append({'name': name, 'stu_id': sid, 'parent_phone': phone, 'status': status})
            elif filename.endswith('.xls') or (xlrd and not filename.endswith('.xlsx')):
                # old Excel .xls；若是文本伪装成 .xls 也尝试降级为 CSV/TSV 解析
                if not xlrd:
                    return err('.xls import requires xlrd', http_status=500)
                rows_iter = None
                try:
                    book = xlrd.open_workbook(file_contents=content)
                    sheet = book.sheet_by_index(0)
                    rows_iter = [sheet.row_values(i) for i in range(sheet.nrows)]
                except Exception:
                    rows_iter = None

                if rows_iter:
                    header, header_idx = _find_header_row(rows_iter)
                    if header is None:
                        data = {'parsed_header': rows_iter[0]} if _debug_headers_allowed() else None
                        return err('invalid header: expected name and stu_id (or 姓名 and 学号)', http_status=400, data=data)
                    if header_idx and header_idx > 0:
                        rows_iter = rows_iter[header_idx:]
                    def find_idx(keys):
                        for k in keys:
                            if k in header:
                                return header.index(k)
                        return None
                    idx_name = find_idx(['name','姓名'])
                    idx_sid = find_idx(['stu_id','学号'])
                    idx_phone = find_idx(['parent_phone','家长电话'])
                    idx_status = find_idx(['status','状态'])
                    rows = []
                    for row in rows_iter[1:]:
                        name = str(row[idx_name]).strip() if (idx_name is not None and idx_name < len(row) and row[idx_name] is not None) else ''
                        sid = str(row[idx_sid]).strip() if (idx_sid is not None and idx_sid < len(row) and row[idx_sid] is not None) else ''
                        phone = str(row[idx_phone]).strip() if (idx_phone is not None and idx_phone < len(row) and row[idx_phone] is not None) else ''
                        status = str(row[idx_status]).strip() if (idx_status is not None and idx_status < len(row) and row[idx_status] is not None) else 'joined'
                        if not name:
                            continue
                        rows.append({'name': name, 'stu_id': sid, 'parent_phone': phone, 'status': status})
                else:
                    # fallback: treat as text (csv/tsv) misnamed as .xls
                    try:
                        txt = content.decode('utf-8', errors='ignore')
                        if not txt.strip():
                            return err('invalid xls file', http_status=400)
                        first_line = txt.splitlines()[0] if txt.splitlines() else ''
                        delimiter = '\t' if '\t' in first_line else ','
                        reader = csv.reader(io.StringIO(txt), delimiter=delimiter)
                        lines = [r for r in reader if any(cell.strip() for cell in r)]
                        if not lines:
                            return err('invalid xls file', http_status=400)
                        header, header_idx = _find_header_row(lines)
                        if header is None:
                            data = {'parsed_header': lines[0]} if _debug_headers_allowed() else None
                            return err('invalid header: expected name and stu_id (or 姓名 and 学号)', http_status=400, data=data)
                        if header_idx and header_idx > 0:
                            lines = lines[header_idx:]
                        try:
                            idx_name = header.index('name') if 'name' in header else (header.index('姓名') if '姓名' in header else None)
                        except ValueError:
                            idx_name = None
                        try:
                            idx_sid = header.index('stu_id') if 'stu_id' in header else (header.index('学号') if '学号' in header else None)
                        except ValueError:
                            idx_sid = None
                        try:
                            idx_phone = header.index('parent_phone') if 'parent_phone' in header else (header.index('家长电话') if '家长电话' in header else None)
                        except ValueError:
                            idx_phone = None
                        try:
                            idx_status = header.index('status') if 'status' in header else (header.index('状态') if '状态' in header else None)
                        except ValueError:
                            idx_status = None
                        rows = []
                        for r in lines[1:]:
                            name = r[idx_name].strip() if (idx_name is not None and idx_name < len(r)) else ''
                            sid = r[idx_sid].strip() if (idx_sid is not None and idx_sid < len(r)) else ''
                            phone = r[idx_phone].strip() if (idx_phone is not None and idx_phone < len(r)) else ''
                            status = r[idx_status].strip() if (idx_status is not None and idx_status < len(r)) else 'joined'
                            if not name:
                                continue
                            rows.append({'name': name, 'stu_id': sid, 'parent_phone': phone, 'status': status})
                    except Exception:
                        return err('invalid xls file', http_status=400)
            else:
                # assume CSV
                txt = content.decode('utf-8', errors='ignore')
                reader = csv.reader(io.StringIO(txt))
                lines = [r for r in reader if any(cell.strip() for cell in r)]
                if not lines:
                    return err('empty csv', http_status=400)
                header, header_idx = _find_header_row(lines)
                if header is None:
                    data = {'parsed_header': lines[0]} if _debug_headers_allowed() else None
                    return err('invalid header: expected name and stu_id (or 姓名 and 学号)', http_status=400, data=data)
                if header_idx and header_idx > 0:
                    lines = lines[header_idx:]
                try:
                    idx_name = header.index('name') if 'name' in header else (header.index('姓名') if '姓名' in header else None)
                except ValueError:
                    idx_name = None
                try:
                    idx_sid = header.index('stu_id') if 'stu_id' in header else (header.index('学号') if '学号' in header else None)
                except ValueError:
                    idx_sid = None
                try:
                    idx_phone = header.index('parent_phone') if 'parent_phone' in header else (header.index('家长电话') if '家长电话' in header else None)
                except ValueError:
                    idx_phone = None
                try:
                    idx_status = header.index('status') if 'status' in header else (header.index('状态') if '状态' in header else None)
                except ValueError:
                    idx_status = None
                rows = []
                for r in lines[1:]:
                    name = r[idx_name].strip() if (idx_name is not None and idx_name < len(r)) else ''
                    sid = r[idx_sid].strip() if (idx_sid is not None and idx_sid < len(r)) else ''
                    phone = r[idx_phone].strip() if (idx_phone is not None and idx_phone < len(r)) else ''
                    status = r[idx_status].strip() if (idx_status is not None and idx_status < len(r)) else 'joined'
                    if not name:
                        continue
                    rows.append({'name': name, 'stu_id': sid, 'parent_phone': phone, 'status': status})
        else:
            # try to read CSV from raw body or form fallback
            raw = ''
            if request.form.get('csv'):
                raw = request.form.get('csv')
            else:
                raw = (request.data or b'').decode('utf-8', errors='ignore')
            if not raw:
                return err('no students data provided', http_status=400)
            reader = csv.reader(io.StringIO(raw))
            lines = [r for r in reader if any(cell.strip() for cell in r)]
            if not lines:
                return err('empty csv', http_status=400)
            header = [_normalize_header(h) for h in lines[0]]
            hk = set(header)
            if not (('name' in hk or '姓名' in hk) and ('stu_id' in hk or '学号' in hk)):
                data = {'parsed_header': header} if _debug_headers_allowed() else None
                return err('invalid header: expected name and stu_id (or 姓名 and 学号)', http_status=400, data=data)
            # find indexes
            try:
                idx_name = header.index('name') if 'name' in header else (header.index('姓名') if '姓名' in header else None)
            except ValueError:
                idx_name = None
            try:
                idx_sid = header.index('stu_id') if 'stu_id' in header else (header.index('学号') if '学号' in header else None)
            except ValueError:
                idx_sid = None
            try:
                idx_phone = header.index('parent_phone') if 'parent_phone' in header else (header.index('家长电话') if '家长电话' in header else None)
            except ValueError:
                idx_phone = None
            try:
                idx_status = header.index('status') if 'status' in header else (header.index('状态') if '状态' in header else None)
            except ValueError:
                idx_status = None

            rows = []
            for r in lines[1:]:
                name = r[idx_name].strip() if (idx_name is not None and idx_name < len(r)) else ''
                sid = r[idx_sid].strip() if (idx_sid is not None and idx_sid < len(r)) else ''
                phone = r[idx_phone].strip() if (idx_phone is not None and idx_phone < len(r)) else ''
                status = r[idx_status].strip() if (idx_status is not None and idx_status < len(r)) else 'joined'
                if not name:
                    continue
                rows.append({'name': name, 'stu_id': sid, 'parent_phone': phone, 'status': status})

    # process rows
    for item in rows:
        name = (item.get('name') or '').strip()
        if not name:
            continue
        stu_id = (item.get('stu_id') or '').strip()
        status = (item.get('status') or 'joined')
        parent_phone = item.get('parent_phone') or ''
        accuracy = item.get('accuracy') if 'accuracy' in item else None
        submit = item.get('submit') if 'submit' in item else None

        existing = None
        if stu_id:
            existing = Student.query.filter_by(class_id=cid, stu_id=stu_id).first()
        if existing:
            existing.name = name
            existing.status = status
            existing.parent_phone = parent_phone
            if accuracy is not None:
                try:
                    existing.accuracy = int(accuracy)
                except Exception:
                    pass
            if submit is not None:
                try:
                    existing.submit = int(submit)
                except Exception:
                    pass
            updated += 1
        else:
            s = Student(
                name=name,
                stu_id=stu_id,
                status=status,
                parent_phone=parent_phone,
                accuracy=(int(accuracy) if (accuracy is not None and str(accuracy).strip() != '') else None),
                submit=(int(submit) if (submit is not None and str(submit).strip() != '') else None),
                class_id=cid
            )
            db.session.add(s)
            added += 1

    db.session.commit()
    # return summary and updated class data
    return ok({'added': added, 'updated': updated, 'class': c.to_dict(include_students=True)}, 'import complete')


@bp.route('/<int:cid>/stats', methods=['GET'])
def class_stats(cid: int):
    uid = _get_uid()
    if not uid:
        return err('missing X-User-Id', http_status=401)
    c = Classroom.query.get(cid)
    if not c or c.created_by != uid:
        return err('class not found', http_status=404)

    students = c.students or []
    total = len(students)
    submitted = sum(1 for s in students if (s.submit is not None and s.submit > 0))
    acc_values = [s.accuracy for s in students if s.accuracy is not None]
    avg_acc = (sum(acc_values) / len(acc_values)) if acc_values else None
    pending = sum(1 for s in students if s.status == 'pending')

    return ok({'total': total, 'submitted': submitted, 'avg_accuracy': (round(avg_acc,2) if avg_acc is not None else None), 'pending': pending})
