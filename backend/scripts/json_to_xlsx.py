import json
import openpyxl
from openpyxl.utils import get_column_letter

inpath = 'backend/exports/class_2.json'
outpath = 'backend/exports/class_2_students.xlsx'

with open(inpath, 'r', encoding='utf-8') as f:
    data = json.load(f)

students = data.get('students', [])
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Students'
headers = ['id','name','stu_id','status','parent_phone','accuracy','submit','created_at']
ws.append(headers)
for s in students:
    ws.append([
        s.get('id',''),
        s.get('name',''),
        s.get('stu_id',''),
        s.get('status',''),
        s.get('parent_phone',''),
        s.get('accuracy') if s.get('accuracy') is not None else '',
        s.get('submit') if s.get('submit') is not None else '',
        s.get('created_at','')
    ])
# auto width
for i, col in enumerate(ws.columns, 1):
    max_length = 0
    for cell in col:
        try:
            v = str(cell.value or '')
        except Exception:
            v = ''
        if len(v) > max_length:
            max_length = len(v)
    ws.column_dimensions[get_column_letter(i)].width = min(50, max(10, max_length + 2))

wb.save(outpath)
print('saved', outpath)
